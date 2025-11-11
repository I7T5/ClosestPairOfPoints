# usr/bin/env python3
"""
Prerequisite: `pip install numpy` in this directory in your console/terminal if you don't have it already.
Run `python3 algorithm.py` to test the closest pair algorithms implemented below.
"""

__author__ = "Alex Bennet, Ella Kocher, Yina Tang"
__description__ = "Implementation of closest pair algorithms. (Project 1 for COMP 422, Fall 2025)"

import argparse
from time import time
from math import isclose
import numpy as np  # run `pip install numpy` in this directory in your console/terminal if you don't have it already
from numpy.typing import NDArray


def brute_force(points: NDArray) -> tuple[tuple[tuple[float, float]], float]:
    """Compute the minimum distance between points using brute-force approach.

    Arguments
    ---------
        points: NDArray
            An array of shape (n, 2) where n is the number of points and 2 is the dimension.

    Returns
    -------
        closest_points: tuple[tuple[float, float]])
            An array of shape (2, 2), i.e., the pair of points with the minimum distance.
        min_dist: float
            The minimum distance between the closest pair of points.
    """
    min_dist = float('inf')
    closest_points = None
    n = points.shape[0]
    
    for i in range(n):
        for j in range(i + 1, n):
            dist = np.linalg.norm(points[i] - points[j])
            if dist < min_dist:
                min_dist = dist
                closest_points = (points[i].tolist(), points[j].tolist())

    return closest_points, min_dist


def divide_and_conquer(points: NDArray) -> tuple[tuple[tuple[float]], float]:
    """Compute the minimum distance between points using an optimized approach. Referenced the pseudocode from [GeeksForGeeks](https://www.geeksforgeeks.org/dsa/closest-pair-of-points-using-divide-and-conquer-algorithm/). 

    Arguments
    ---------
        points: NDArray
            An array of shape (n, 2) where n is the number of points and 2 is the dimension.

    Returns
    -------
        points: tuple[tuple[float, float]]
            An array of shape (2, d), i.e., the pair of points with the minimum distance.
        distance: float
            The minimum distance between the closest pair of points.
    """

    # Steps: 
    # 1. Sort points by x-coordinate
    # 2. Recursively divide the set of points into two halves
    # 3. Find the closest pair in each half
    # 4. Find the closest pair across the dividing line
    # 5. Return the overall closest pair

    def recurse(A: NDArray, A_y: NDArray) -> tuple[tuple[tuple[float]], float]:
        """Main recursive function for divide and conquer approach.

        Parameters
        ----------
        A : NDArray
            NDArray of size n times 2 points (tuple of two floats) in 2D sorted by x-coordinate
        A_y : NDArray
            NDArray of size n times 2 points (tuple of two floats) in 2D sorted by y-coordinate

        Returns
        -------
        tuple[tuple[tuple[float]], float]
            tuple[tuple[float]]: Closest pair of points. A pair of tuples, each with of two floats
            float: Distance between the closest pair of points
        """
        # Base cases
        if len(A) <= 1:
            return (([float('-inf'), float('-inf')], [float('inf'), float('inf')]), float('inf'))
        elif len(A) == 2:
            return (A.tolist(), float(np.linalg.norm(A[0] - A[1])))
        
        # Recursive case
        # Find the midpoint i.e. pivot
        mid_i = len(A) // 2

        # Divide A_y into left and right halves by x-coordinate, preserving the y-ordering
        A_y_left = []
        A_y_right = []
        for i in range(len(A_y)):
            point = A_y[i]
            if point[0] <= A[mid_i-1, 0]:
                A_y_left.append(point)
            else:
                A_y_right.append(point)
            
            # Stop once we've filled the left half; ensures |A| == |A_y|
            if len(A_y_left) == mid_i:
                if (i+1) < (len(A_y)-1):
                    A_y_right.extend(A_y[i+1:])
                break

        # 2. Recursively divide the set of points into two halves
        cp_left, d_left = recurse(A[:mid_i], A_y_left)
        cp_right, d_right = recurse(A[mid_i:], A_y_right)

        # 3. Find the closest pair in each half
        if d_left < d_right:
            cp = cp_left
            d = d_left
        else:
            cp = cp_right
            d = d_right

        # 4. Find the closest pair across the dividing line
        # Collect points whose x-distance from the midline is ≤ d. Call this the "strip."
        # Collect points one-by-one from A_y so the strip is sorted by y-coordinate
        strip = []
        for points in A_y:
            if abs(points[0] - A[mid_i][0]) <= d:
                strip.append(points)

        # For each point in the strip, compare it with the next up to 7 points (having y-distance ≤ d) to check for closer pairs.
        for i in range(len(strip)):
            stop = min(i + 8, len(strip))
            for j in range(i+1, stop):
                if (strip[j][1] - strip[i][1]) > d:
                    break

                dist = np.linalg.norm(strip[i] - strip[j])
                if dist < d:
                    cp = (strip[i].tolist(), strip[j].tolist())
                    d = dist

        # 5. Return the overall closest pair
        return cp, float(d)

    # 1. Sort points by x-coordinate and y-coordinate, respectively
    A = points[points[:, 0].argsort()]
    A_y = points[points[:, 1].argsort()]

    # 2. Recursively divide the set of points into two halves
    return recurse(A, A_y)


def main(output_file: str) -> None:
    """Collect timing data for both algorithms and save to Excel file.
    
    Tests both brute force and divide-and-conquer algorithms with input sizes
    n = 1, 50, 100, 200, 400, 800, 1500, 2000, 3000, running 10 trials for each size.
    
    Arguments
    ---------
        output_file: str
            The name of the Excel file to save results to.
    
    Usage Examples
    --------------
    Run with default collect comprehensive timing data for analysis (runs multiple trials):
        python algorithm.py
    
    Collect timing data with custom output filename:
        python algorithm.py --collect-data --output-file my_results.xlsx
        python algorithm.py -cd -o my_results.xlsx
    """
    from openpyxl import Workbook # run `pip install openpyxl` in this directory in your console/terminal to get the module
    from openpyxl.styles import Font, PatternFill, Alignment
    
    test_sizes = [1, 50, 100, 200, 400, 800, 1500, 2000, 3000]
    num_trials = 10
    
    # Store results
    brute_force_times = {n: [] for n in test_sizes}
    divide_conquer_times = {n: [] for n in test_sizes}
    
    print("Collecting timing data...")
    
    for n in test_sizes:
        print(f"\nTesting with n={n} points...")
        
        for trial in range(num_trials):
            # Generate random points
            coords = np.random.uniform(-100, 100, n * 2)
            points = coords.reshape((n, 2))
            
            # Time brute force
            start = time()
            brute_force(points)
            bf_time = time() - start
            brute_force_times[n].append(bf_time)
            
            # Time divide and conquer
            start = time()
            divide_and_conquer(points)
            dc_time = time() - start
            divide_conquer_times[n].append(dc_time)
            
            print(f"  Trial {trial + 1}/10: BF={bf_time:.6f}s, D&C={dc_time:.6f}s")
    
    print(f"\nSaving results to {output_file}...")
    
    # Create Excel workbook
    wb = Workbook()
    
    # Sheet 1: Brute Force Data
    bf_sheet = wb.active
    bf_sheet.title = "Brute Force"
    
    # Headers
    bf_sheet['A1'] = 'Input Size (n)'
    bf_sheet['A1'].font = Font(bold=True)
    for i in range(num_trials):
        bf_sheet.cell(1, i + 2, f'Trial {i + 1}')
        bf_sheet.cell(1, i + 2).font = Font(bold=True)
    bf_sheet.cell(1, num_trials + 2, 'Mean')
    bf_sheet.cell(1, num_trials + 2).font = Font(bold=True, color='0000FF')
    bf_sheet.cell(1, num_trials + 3, 'Std Dev')
    bf_sheet.cell(1, num_trials + 3).font = Font(bold=True, color='0000FF')
    
    # Data
    for idx, n in enumerate(test_sizes, start=2):
        bf_sheet.cell(idx, 1, n)
        for trial_idx, time_val in enumerate(brute_force_times[n], start=2):
            bf_sheet.cell(idx, trial_idx, time_val)
        
        # Formulas for mean and std dev
        bf_sheet.cell(idx, num_trials + 2, f'=AVERAGE(B{idx}:{chr(65 + num_trials)}{idx})')
        bf_sheet.cell(idx, num_trials + 3, f'=STDEV.S(B{idx}:{chr(65 + num_trials)}{idx})')
    
    # Formatting
    bf_sheet.column_dimensions['A'].width = 15
    for col in range(2, num_trials + 4):
        bf_sheet.column_dimensions[chr(64 + col)].width = 12
    
    # Sheet 2: Divide and Conquer Data
    dc_sheet = wb.create_sheet("Divide and Conquer")
    
    # Headers
    dc_sheet['A1'] = 'Input Size (n)'
    dc_sheet['A1'].font = Font(bold=True)
    for i in range(num_trials):
        dc_sheet.cell(1, i + 2, f'Trial {i + 1}')
        dc_sheet.cell(1, i + 2).font = Font(bold=True)
    dc_sheet.cell(1, num_trials + 2, 'Mean')
    dc_sheet.cell(1, num_trials + 2).font = Font(bold=True, color='0000FF')
    dc_sheet.cell(1, num_trials + 3, 'Std Dev')
    dc_sheet.cell(1, num_trials + 3).font = Font(bold=True, color='0000FF')
    
    # Data
    for idx, n in enumerate(test_sizes, start=2):
        dc_sheet.cell(idx, 1, n)
        for trial_idx, time_val in enumerate(divide_conquer_times[n], start=2):
            dc_sheet.cell(idx, trial_idx, time_val)
        
        # Formulas for mean and std dev
        dc_sheet.cell(idx, num_trials + 2, f'=AVERAGE(B{idx}:{chr(65 + num_trials)}{idx})')
        dc_sheet.cell(idx, num_trials + 3, f'=STDEV.S(B{idx}:{chr(65 + num_trials)}{idx})')
    
    # Formatting
    dc_sheet.column_dimensions['A'].width = 15
    for col in range(2, num_trials + 4):
        dc_sheet.column_dimensions[chr(64 + col)].width = 12
    
    # Sheet 3: Summary
    summary_sheet = wb.create_sheet("Summary")
    
    # Headers
    summary_sheet['A1'] = 'Input Size (n)'
    summary_sheet['B1'] = 'Brute Force Mean (s)'
    summary_sheet['C1'] = 'Brute Force Std Dev (s)'
    summary_sheet['D1'] = 'D&C Mean (s)'
    summary_sheet['E1'] = 'D&C Std Dev (s)'
    summary_sheet['F1'] = 'Speedup Factor'
    
    for cell in ['A1', 'B1', 'C1', 'D1', 'E1', 'F1']:
        summary_sheet[cell].font = Font(bold=True)
        summary_sheet[cell].fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
    
    # Data with formulas referencing other sheets
    for idx, n in enumerate(test_sizes, start=2):
        summary_sheet.cell(idx, 1, n)
        summary_sheet.cell(idx, 2, f"='Brute Force'!L{idx}")
        summary_sheet.cell(idx, 3, f"='Brute Force'!M{idx}")
        summary_sheet.cell(idx, 4, f"='Divide and Conquer'!L{idx}")
        summary_sheet.cell(idx, 5, f"='Divide and Conquer'!M{idx}")
        summary_sheet.cell(idx, 6, f"=B{idx}/D{idx}")
    
    # Formatting
    summary_sheet.column_dimensions['A'].width = 15
    for col in ['B', 'C', 'D', 'E', 'F']:
        summary_sheet.column_dimensions[col].width = 20
    
    # Save workbook
    wb.save(output_file)
    print(f"Results saved successfully to {output_file}")
    print("\nSummary:")
    print(f"  Test sizes: {test_sizes}")
    print(f"  Trials per size: {num_trials}")
    print(f"  Total tests run: {len(test_sizes) * num_trials * 2}")


if __name__ == "__main__":
    # Parse command-line arguments
    parser = argparse.ArgumentParser(description="Closest pair of points algorithms.")

    # Deprecated options commented out
    # parser.add_argument("--num-points", "-np", type=int, help="number of points", default=5000)
    # parser.add_argument("--collect-data", "-cd", action="store_true", help="collect timing data and save to Excel")

    parser.add_argument("--output-file", "-o", type=str, default="timing_results.xlsx", help="output Excel file name")
    args = parser.parse_args()
    
    main(args.output_file)